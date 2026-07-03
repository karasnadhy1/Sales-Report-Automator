"""
Shared pipeline logic used by app.py.

This is split out from the app so the cleaning/report logic can be
tested on its own, and so app.py stays focused on just the UI.

Key design change from Day 1-3: instead of assuming fixed column
names, we now DETECT which column in the client's file matches each
field we need (date, price, quantity, etc.) using keyword matching,
then let the user confirm/fix the guess before processing.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# Fields we look for, and the keywords that suggest a column matches them.
# Order matters: first keyword match wins.
FIELD_KEYWORDS = {
    "order_date":   ["date", "order date", "purchase date"],
    "quantity":     ["quantity", "qty", "units", "count"],
    "unit_price":   ["price", "unit price", "cost", "amount"],
    "product":      ["product", "item", "sku"],
    "category":     ["category", "type", "department"],
    "region":       ["region", "state", "territory", "location"],
    "sales_rep":    ["sales rep", "rep", "salesperson", "agent"],
    "customer_name": ["customer", "client", "buyer", "name"],
}
REQUIRED_FIELDS = ["order_date", "quantity", "unit_price"]
OPTIONAL_FIELDS = ["product", "category", "region", "sales_rep", "customer_name"]


def suggest_column_mapping(df: pd.DataFrame) -> dict:
    """Best-guess mapping of {field_name: actual_column_name_or_None}."""
    columns_lower = {c: c.lower().strip() for c in df.columns}
    mapping = {}
    for field, keywords in FIELD_KEYWORDS.items():
        match = None
        for col, col_lower in columns_lower.items():
            if any(kw in col_lower for kw in keywords):
                match = col
                break
        mapping[field] = match
    return mapping


def generate_sample_csv() -> pd.DataFrame:
    """
    A small, realistically messy sample dataset so a first-time visitor
    can try the app with one click, with no CSV of their own on hand.
    Uses real randomness (no fixed seed) so every click produces a
    genuinely different dataset — a repeat visitor sees something new,
    not the exact same numbers every time.
    """
    import random
    customers = ["Maria Santos", "James O'Connor", "Liu Wei", "Tom Becker", "Priya Nair"]
    products = {"Laptop": ("Electronics", 899.99), "Desk Lamp": ("Home", 24.50),
                "Notebook": ("Office", 4.50), "Headphones": ("Electronics", 59.99)}
    regions = ["North", "South", "East", "West"]
    rows = []
    for i in range(1, 41):
        product = random.choice(list(products.keys()))
        category, price = products[product]
        qty = random.randint(1, 5) * random.choice([1, 1, 1, -1])  # occasional return
        rows.append({
            "order_id": 2000 + i,
            "order_date": f"{random.randint(1,12)}/{random.randint(1,28)}/2026",
            "customer_name": random.choice(customers),
            "product": product,
            "category": random.choice([category, category.lower()]),
            "quantity": qty,
            "unit_price": f"${price}" if random.random() < 0.3 else price,
            "region": random.choice(regions),
            "sales_rep": random.choice(["A. Kim", "B. Ortiz"]),
        })
    return pd.DataFrame(rows)


def clean_dataframe(raw_df: pd.DataFrame, mapping: dict) -> tuple[pd.DataFrame, dict]:
    """
    Standardizes a raw dataframe into fixed internal column names based
    on the user-confirmed mapping, then applies the same cleaning steps
    from Day 1: whitespace stripping, casing, date parsing, price/qty
    conversion, duplicate removal.

    Returns (clean_df, diagnostics) where diagnostics explains exactly
    what was removed and why — never drop rows silently.
    """
    df = pd.DataFrame()
    diagnostics = {"starting_rows": len(raw_df)}

    for field in REQUIRED_FIELDS:
        source_col = mapping.get(field)
        if source_col is None or source_col not in raw_df.columns:
            raise ValueError(f"Required field '{field}' is not mapped to a column.")
        df[field] = raw_df[source_col].astype(str)

    for field in OPTIONAL_FIELDS:
        source_col = mapping.get(field)
        if source_col is not None and source_col in raw_df.columns:
            df[field] = raw_df[source_col].astype(str)

    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = df[col].str.strip()

    for col in ("category", "region"):
        if col in df.columns:
            df[col] = df[col].str.title()

    def parse_date(value):
        if pd.isna(value) or value == "" or value.lower() == "nan":
            return pd.NaT
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                return pd.to_datetime(value, format=fmt)
            except ValueError:
                continue
        try:
            return pd.to_datetime(value)
        except (ValueError, TypeError):
            return pd.NaT

    df["order_date"] = df["order_date"].apply(parse_date)
    df["unit_price"] = (
        df["unit_price"]
        .str.replace(r"[^\d.\-]", "", regex=True)
        .replace("", pd.NA)
        .astype(float)
    )
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce")
    df["is_return"] = df["quantity"] < 0
    df["quantity"] = df["quantity"].abs()

    # --- Diagnostics: count exactly what's wrong BEFORE dropping anything ---
    diagnostics["missing_date"] = int(df["order_date"].isna().sum())
    diagnostics["missing_price"] = int(df["unit_price"].isna().sum())
    diagnostics["missing_quantity"] = int(df["quantity"].isna().sum())

    before_dedup = len(df)
    df = df.drop_duplicates()
    diagnostics["duplicates_removed"] = before_dedup - len(df)

    before_dropna = len(df)
    df = df.dropna(subset=["order_date", "unit_price", "quantity"])
    diagnostics["unrecoverable_removed"] = before_dropna - len(df)
    diagnostics["ending_rows"] = len(df)

    for col in ("category", "region", "sales_rep", "customer_name", "product"):
        if col in df.columns:
            df[col] = df[col].replace("", "Unspecified").fillna("Unspecified")

    df["revenue"] = df["quantity"] * df["unit_price"]
    if len(df) > 0:
        df["month"] = df["order_date"].dt.strftime("%Y-%m")
        df = df.sort_values("order_date").reset_index(drop=True)
    return df, diagnostics


def build_excel_report(df: pd.DataFrame) -> bytes:
    """
    Builds the same style of formula-driven Excel report as Day 3,
    but adapts which tables/charts to include based on which optional
    fields actually exist in this dataset.
    """
    FONT_NAME = "Arial"
    HEADER_FILL = PatternFill("solid", start_color="1F4E78")
    HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
    LABEL_FONT = Font(name=FONT_NAME, bold=True, size=11)
    BODY_FONT = Font(name=FONT_NAME, size=11)
    CURRENCY_FMT = '$#,##0.00'

    wb = Workbook()
    data_sheet = wb.active
    data_sheet.title = "Data"

    columns = list(df.columns)
    col_index = {name: i + 1 for i, name in enumerate(columns)}  # 1-based

    for col_idx, col_name in enumerate(columns, start=1):
        cell = data_sheet.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = data_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            colname = columns[col_idx - 1]
            if colname in ("unit_price", "revenue"):
                cell.number_format = CURRENCY_FMT
            if colname == "order_date":
                cell.number_format = "yyyy-mm-dd"

    last_row = len(df) + 1
    for col_idx in range(1, len(columns) + 1):
        data_sheet.column_dimensions[get_column_letter(col_idx)].width = 16
    data_sheet.freeze_panes = "A2"

    summary = wb.create_sheet("Summary")
    summary.sheet_view.showGridLines = False
    summary["B2"] = "Sales Report"
    summary["B2"].font = TITLE_FONT
    summary["B3"] = f"Generated from {len(df)} orders"
    summary["B3"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")

    is_return_col = get_column_letter(col_index["is_return"])
    revenue_col = get_column_letter(col_index["revenue"])

    kpis = [
        ("Total Revenue", f'=SUMIF(Data!{is_return_col}2:{is_return_col}{last_row},FALSE,'
                           f'Data!{revenue_col}2:{revenue_col}{last_row})', CURRENCY_FMT),
        ("Total Orders", f'=COUNTA(Data!A2:A{last_row})', '#,##0'),
        ("Returns (value)", f'=SUMIF(Data!{is_return_col}2:{is_return_col}{last_row},TRUE,'
                             f'Data!{revenue_col}2:{revenue_col}{last_row})', CURRENCY_FMT),
    ]
    for (label, formula, fmt), col in zip(kpis, ["B", "D", "F"]):
        summary[f"{col}5"] = label
        summary[f"{col}5"].font = LABEL_FONT
        summary[f"{col}6"] = formula
        summary[f"{col}6"].font = Font(name=FONT_NAME, size=14, bold=True, color="1F4E78")
        summary[f"{col}6"].number_format = fmt

    next_row = 9
    chart_anchor_row = 5

    def add_breakdown_table(field, title, start_col_letter, row_start):
        if field not in df.columns:
            return None
        values = sorted(df[field].dropna().unique())
        col_letter_a, col_letter_b = start_col_letter, chr(ord(start_col_letter) + 1)
        summary[f"{col_letter_a}{row_start - 1}"] = title
        summary[f"{col_letter_a}{row_start - 1}"].font = LABEL_FONT
        summary[f"{col_letter_a}{row_start}"] = title.split(" by ")[-1]
        summary[f"{col_letter_b}{row_start}"] = "Revenue"
        summary[f"{col_letter_a}{row_start}"].font = HEADER_FONT
        summary[f"{col_letter_a}{row_start}"].fill = HEADER_FILL
        summary[f"{col_letter_b}{row_start}"].font = HEADER_FONT
        summary[f"{col_letter_b}{row_start}"].fill = HEADER_FILL
        field_col = get_column_letter(col_index[field])
        for i, val in enumerate(values):
            r = row_start + 1 + i
            summary[f"{col_letter_a}{r}"] = val
            summary[f"{col_letter_a}{r}"].font = BODY_FONT
            summary[f"{col_letter_b}{r}"] = (
                f'=SUMIFS(Data!{revenue_col}2:{revenue_col}{last_row},'
                f'Data!{field_col}2:{field_col}{last_row},{col_letter_a}{r},'
                f'Data!{is_return_col}2:{is_return_col}{last_row},FALSE)'
            )
            summary[f"{col_letter_b}{r}"].number_format = CURRENCY_FMT
            summary[f"{col_letter_b}{r}"].font = BODY_FONT
        end_row = row_start + len(values)
        summary.conditional_formatting.add(
            f"{col_letter_b}{row_start + 1}:{col_letter_b}{end_row}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           end_type="max", end_color="63BE7B")
        )
        return row_start, end_row, col_letter_a, col_letter_b

    table_positions = {}
    col_cycle = ["B", "E"]
    ci = 0
    for field, title in [("category", "Revenue by Category"), ("region", "Revenue by Region")]:
        if field in df.columns:
            pos = add_breakdown_table(field, title, col_cycle[ci % 2], next_row)
            if pos:
                table_positions[field] = pos
                ci += 1

    max_end_row = max((p[1] for p in table_positions.values()), default=next_row)
    month_start = max_end_row + 3
    months = sorted(df["month"].unique())
    summary[f"B{month_start - 1}"] = "Monthly Revenue"
    summary[f"B{month_start - 1}"].font = LABEL_FONT
    summary[f"B{month_start}"] = "Month"
    summary[f"C{month_start}"] = "Revenue"
    summary[f"B{month_start}"].font = summary[f"C{month_start}"].font = HEADER_FONT
    summary[f"B{month_start}"].fill = summary[f"C{month_start}"].fill = HEADER_FILL
    month_col = get_column_letter(col_index["month"])
    for i, month in enumerate(months):
        r = month_start + 1 + i
        summary[f"B{r}"] = month
        summary[f"B{r}"].font = BODY_FONT
        summary[f"C{r}"] = (
            f'=SUMIFS(Data!{revenue_col}2:{revenue_col}{last_row},'
            f'Data!{month_col}2:{month_col}{last_row},B{r},'
            f'Data!{is_return_col}2:{is_return_col}{last_row},FALSE)'
        )
        summary[f"C{r}"].number_format = CURRENCY_FMT
        summary[f"C{r}"].font = BODY_FONT
    month_end = month_start + len(months)

    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        summary.column_dimensions[col].width = 18

    # Charts: one per breakdown table that exists, plus the monthly trend
    chart_row = 5
    for field, (row_start, end_row, col_a, col_b) in table_positions.items():
        bar = BarChart()
        bar.title = f"Revenue by {field.title()}"
        bar.y_axis.title = "Revenue ($)"
        bar.style = 10
        col_a_idx = ord(col_a) - ord("A") + 1
        col_b_idx = ord(col_b) - ord("A") + 1
        data_ref = Reference(summary, min_col=col_b_idx, min_row=row_start, max_row=end_row)
        cat_ref = Reference(summary, min_col=col_a_idx, min_row=row_start + 1, max_row=end_row)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cat_ref)
        bar.width, bar.height = 14, 8
        summary.add_chart(bar, f"J{chart_row}")
        chart_row += 17

    line = LineChart()
    line.title = "Monthly Revenue Trend"
    line.y_axis.title = "Revenue ($)"
    line.style = 12
    month_data = Reference(summary, min_col=3, min_row=month_start, max_row=month_end)
    month_labels = Reference(summary, min_col=2, min_row=month_start + 1, max_row=month_end)
    line.add_data(month_data, titles_from_data=True)
    line.set_categories(month_labels)
    line.width, line.height = 14, 8
    summary.add_chart(line, f"J{chart_row}")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
