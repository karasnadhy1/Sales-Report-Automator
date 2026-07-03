"""
Day 3: Turn clean_sales_data.csv into a polished, client-ready Excel
report — a Data sheet plus a Summary sheet with KPIs, breakdown
tables, charts, and conditional formatting.

Key design choice: the Summary sheet uses REAL EXCEL FORMULAS
(SUM, SUMIF, COUNTA, AVERAGE) that reference the Data sheet, instead
of pasting in numbers we calculated in Python. This means if a client
pastes in a new month of data on the Data sheet, the whole report
recalculates itself — no need to re-run Python. That's the difference
between a one-off spreadsheet and a genuinely reusable tool.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=11)
BODY_FONT = Font(name=FONT_NAME, size=11)
CURRENCY_FMT = '$#,##0.00'
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))

df = pd.read_csv("clean_sales_data.csv", parse_dates=["order_date"])
df["month"] = df["order_date"].dt.strftime("%Y-%m")

wb = Workbook()

# ============================================================
# SHEET 1: Data (raw clean data the formulas will reference)
# ============================================================
data_sheet = wb.active
data_sheet.title = "Data"

columns = ["order_id", "order_date", "customer_name", "product", "category",
           "quantity", "unit_price", "region", "sales_rep", "is_return", "revenue", "month"]
for col_idx, col_name in enumerate(columns, start=1):
    cell = data_sheet.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

for row_idx, row in enumerate(df[columns].itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = data_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BODY_FONT
        if columns[col_idx - 1] in ("unit_price", "revenue"):
            cell.number_format = CURRENCY_FMT
        if columns[col_idx - 1] == "order_date":
            cell.number_format = "yyyy-mm-dd"

last_row = len(df) + 1
for col_idx, col_name in enumerate(columns, start=1):
    data_sheet.column_dimensions[get_column_letter(col_idx)].width = 16
data_sheet.freeze_panes = "A2"

# ============================================================
# SHEET 2: Summary (formulas + charts + conditional formatting)
# ============================================================
summary = wb.create_sheet("Summary")
summary.sheet_view.showGridLines = False

summary["B2"] = "Monthly Sales Report"
summary["B2"].font = TITLE_FONT
summary["B3"] = f"Generated from {len(df)} orders | {df['order_date'].min().date()} to {df['order_date'].max().date()}"
summary["B3"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")

# --- KPI cards (row 5-6), each a real formula against the Data sheet ---
kpis = [
    ("Total Revenue", f'=SUMIF(Data!J2:J{last_row},FALSE,Data!K2:K{last_row})', CURRENCY_FMT),
    ("Total Orders", f'=COUNTA(Data!A2:A{last_row})', '#,##0'),
    ("Avg Order Value", f'=B6/COUNTIF(Data!J2:J{last_row},FALSE)', CURRENCY_FMT),
    ("Returns (value)", f'=SUMIF(Data!J2:J{last_row},TRUE,Data!K2:K{last_row})', CURRENCY_FMT),
]
kpi_col_starts = ["B", "D", "F", "H"]
for (label, formula, fmt), col in zip(kpis, kpi_col_starts):
    summary[f"{col}5"] = label
    summary[f"{col}5"].font = LABEL_FONT
    summary[f"{col}6"] = formula
    summary[f"{col}6"].font = Font(name=FONT_NAME, size=14, bold=True, color="1F4E78")
    summary[f"{col}6"].number_format = fmt

# --- Revenue by Category table (row 9+) ---
categories = sorted(df["category"].unique())
summary["B8"] = "Revenue by Category"
summary["B8"].font = LABEL_FONT
summary["B9"] = "Category"
summary["C9"] = "Revenue"
summary["B9"].font = summary["C9"].font = HEADER_FONT
summary["B9"].fill = summary["C9"].fill = HEADER_FILL
for i, cat in enumerate(categories):
    r = 10 + i
    summary[f"B{r}"] = cat
    summary[f"B{r}"].font = BODY_FONT
    summary[f"C{r}"] = f'=SUMIFS(Data!K2:K{last_row},Data!E2:E{last_row},B{r},Data!J2:J{last_row},FALSE)'
    summary[f"C{r}"].number_format = CURRENCY_FMT
    summary[f"C{r}"].font = BODY_FONT
cat_end_row = 9 + len(categories)

# Color scale: red (low) -> green (high) so weak categories jump out visually
summary.conditional_formatting.add(
    f"C10:C{cat_end_row}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   end_type="max", end_color="63BE7B")
)

# --- Revenue by Region table (row 9+, columns E-F) ---
regions = sorted(df["region"].unique())
summary["E8"] = "Revenue by Region"
summary["E8"].font = LABEL_FONT
summary["E9"] = "Region"
summary["F9"] = "Revenue"
summary["E9"].font = summary["F9"].font = HEADER_FONT
summary["E9"].fill = summary["F9"].fill = HEADER_FILL
for i, reg in enumerate(regions):
    r = 10 + i
    summary[f"E{r}"] = reg
    summary[f"E{r}"].font = BODY_FONT
    summary[f"F{r}"] = f'=SUMIFS(Data!K2:K{last_row},Data!H2:H{last_row},E{r},Data!J2:J{last_row},FALSE)'
    summary[f"F{r}"].number_format = CURRENCY_FMT
    summary[f"F{r}"].font = BODY_FONT
reg_end_row = 9 + len(regions)
summary.conditional_formatting.add(
    f"F10:F{reg_end_row}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   end_type="max", end_color="63BE7B")
)

# --- Monthly revenue table (row further down, columns B-C) ---
months = sorted(df["month"].unique())
month_table_start = max(cat_end_row, reg_end_row) + 3
summary.cell(row=month_table_start - 1, column=2, value="Monthly Revenue").font = LABEL_FONT
summary.cell(row=month_table_start, column=2, value="Month").font = HEADER_FONT
summary.cell(row=month_table_start, column=2).fill = HEADER_FILL
summary.cell(row=month_table_start, column=3, value="Revenue").font = HEADER_FONT
summary.cell(row=month_table_start, column=3).fill = HEADER_FILL
for i, month in enumerate(months):
    r = month_table_start + 1 + i
    summary.cell(row=r, column=2, value=month).font = BODY_FONT
    formula = f'=SUMIFS(Data!K2:K{last_row},Data!L2:L{last_row},B{r},Data!J2:J{last_row},FALSE)'
    cell = summary.cell(row=r, column=3, value=formula)
    cell.number_format = CURRENCY_FMT
    cell.font = BODY_FONT
month_end_row = month_table_start + len(months)

# --- Column widths ---
for col in ["B", "C", "D", "E", "F", "G", "H"]:
    summary.column_dimensions[col].width = 18

# ============================================================
# Charts
# ============================================================
# Bar chart: revenue by category
bar = BarChart()
bar.title = "Revenue by Category"
bar.y_axis.title = "Revenue ($)"
bar.style = 10
cat_data = Reference(summary, min_col=3, min_row=9, max_row=cat_end_row)
cat_labels = Reference(summary, min_col=2, min_row=10, max_row=cat_end_row)
bar.add_data(cat_data, titles_from_data=True)
bar.set_categories(cat_labels)
bar.width, bar.height = 14, 8
summary.add_chart(bar, "J5")

# Line chart: monthly revenue trend
line = LineChart()
line.title = "Monthly Revenue Trend"
line.y_axis.title = "Revenue ($)"
line.style = 12
month_data = Reference(summary, min_col=3, min_row=month_table_start, max_row=month_end_row)
month_labels = Reference(summary, min_col=2, min_row=month_table_start + 1, max_row=month_end_row)
line.add_data(month_data, titles_from_data=True)
line.set_categories(month_labels)
line.width, line.height = 14, 8
summary.add_chart(line, "J22")

wb.save("sales_report.xlsx")
print("Saved sales_report.xlsx")
print(f"Categories: {categories}")
print(f"Regions: {regions}")
print(f"Months: {months}")
